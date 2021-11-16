from openpyxl import load_workbook
import re
from env import path, filename

## Load workbook ##

wb = load_workbook(f'{path}{filename}')
ws = wb.active

## General Font Formatting ##

# ws.insert_cols(6)
# ws['F1'].value = 'Phone Formatted'
# ws['F1'].font = Font(bold=True, size= 14)


## Format Phone numbers Loop ##

# no_data = 0
# for row in range(1, 840):
#     cv = ws[f'E{row}'].value
#     if not cv:
#         no_data += 1
#         ws[f'F{row}'].value = "None"
#     else:
#         cv = str(cv)
#         new_cell_value = cv[0] +' (' + cv[1:4] + ') ' + cv[4:7] + '-' + cv[7:]
#         ws[f'F{row}'].value = new_cell_value
#
# print (f'No data in {no_data} cells')


## Format clear columns Loop ##

# for row in range(2, 840):
#     ws[f'R{row}'].value = ''


## Pure Python for Interest ##

# for row in range(2,840):
#     comment = ws[f'O{row}'].value
#     if comment is not None:
#         if 'Just browsing' in comment or \
#             'just browsing' in comment or \
#             'Just looking' in comment or \
#             'just looking' in comment:
#             ws[f'Q{row}'].value = 'Just looking'
#         elif 'Buy a home' in comment or \
#             'buy a home?: Yes' in comment or \
#             'Buying a home' in comment or \
#             'buying a home' in comment:
#             ws[f'Q{row}'].value = 'Buy a home'
#         elif 'Sell a home' in comment or \
#             'sell a home?: Yes' in comment or \
#             'selling a home' in comment or \
#             'Thinking of Selling' in comment or \
#             'Thinking about Selling' in comment:
#             ws[f'Q{row}'].value = 'Sell a home'
#         elif 'Looking for Home Evaluation' in comment or \
#             'Property valuation' in comment or \
#             'property valuation' in comment or \
#             'Property Valuation' in comment:
#             ws[f'Q{row}'].value = 'Property valuation'

# Regex for price range:

# for row in range(2, 840):
#     comment = ws[f'O{row}'].value
#     if comment is not None and ws[f'Q{row}'].value == 'Buy a home' or ws[f'Q{row}'].value == 'Just looking':
#         comment_regex = re.compile(r'''
#         (\$?(\d,)?\d\d\d,\d\d\d)+  #Match at least one 3-7 digit number with $ , optional
#         (\s-\s)?                      #Match optional dash separator
#         (\$?(\d,)?\d\d\d,\d\d\d\d?)?''', re.VERBOSE) #Match optional 3-7 digit number with $ , optional
#         mo = comment_regex.search(comment)
#         if mo:
#             if 'under' in comment or \
#                     'Under' in comment or \
#                     'Less' in comment or \
#                     'less' in comment or \
#                     'below' in comment or \
#                     'Below' in comment:
#                 ws[f'R{row}'].value = (f'Less than {mo.group()}')
#             elif 'Max' in comment or \
#                     'max' in comment:
#                 ws[f'R{row}'].value = (f'{mo.group()} maximum')
#             else:
#                 ws[f'R{row}'].value = (mo.group())
#         else:
#             comment_regex = re.compile(
#                 r'(Price range: )(\D*)?((\$?(\d,)?\d\d\dk?,?(\d\d\d)?)+([ to-]*)?(\$?(\d,)?\d\d\dk?,?(\d\d\d\d?)?)?)(.*)*')
#             mo = comment_regex.search(comment)
#             if mo:
#                 if 'under' in mo.group(2) or \
#                         'Under' in mo.group(2) or \
#                         'Less' in mo.group(2) or \
#                         'less' in mo.group(2):
#                     ws[f'R{row}'].value = (f'Less than {mo.group(3)}')
#                 elif 'max' in mo.group(2) or \
#                         'Max' in mo.group(2) or \
#                         'max' in mo.group(4) or \
#                         'max' in comment or \
#                         'Max' in mo.group(4):
#                     ws[f'R{row}'].value = (f'{mo.group(3)} maximum')
#                 else:
#                     ws[f'R{row}'].value = (mo.group(3))

# Pure Python for has agent. Some records show that users were working with but not committed to another agent. So the
# parser only returns yes to agent if the user input matches committed to another agent.

# for row in range(2,840):
#     comment = ws[f'O{row}'].value
#     if comment is not None:
#         if 'Committed to another agent?: Yes' in comment:
#             ws[f'P{row}'].value = 'Yes'
#         else:
#             ws[f'P{row}'].value = 'No'


## Pure python for County of interest ##

# for row in range(2,840):
#     comment = ws[f'O{row}'].value
#     if comment is not None:
#         if 'Suffolk' in comment or 'suffolk' in comment:
#             ws[f'T{row}'].value = 'Suffolk'
#         elif 'Nassau' in comment or 'nassua' in comment:
#             ws[f'T{row}'].value = 'Nassau'
#         elif 'Queens' in comment or 'queens' in comment:
#             ws[f'T{row}'].value = 'Queens'


## Regex for availability ##

# for row in range(2, 840):
#     comment = ws[f'O{row}'].value
#     if comment is not None:
#         if 'Best time to reach' in comment:
#             avail_regex = re.compile(r'(Best time to reach\?: )(.*)')
#             mo = avail_regex.search(comment)
#             ws[f'U{row}'].value = mo.group(2)


## Regex for areas of interest ##

# for row in range(2,840):
#     comment = ws[f'O{row}'].value
#     if comment is not None:
#         areas_regex = re.compile(r'Areas of interest: (.*)')
#         mo = areas_regex.findall(comment)
#         if len(mo) >= 1:
#             ws[f'S{row}'].value = mo[0]


## Regex for Lead assigned to ##

# for row in range(2,840):
#     comment = ws[f'O{row}'].value
#     if comment is not None:
#         assigned_regex = re.compile(r'(Lead was assigned to )(.*)')
#         mo = assigned_regex.search(comment)
#         if mo:
#             ws[f'V{row}'].value = mo.group(2)

## Save File ##

wb.save(filename=f'{path}{filename}')
